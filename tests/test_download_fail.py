import csv
import zipfile
from pypdf import PdfReader
from script_os import ZIP_DIR
from openpyxl import load_workbook


def test_pdf_read():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('file_PDF.pdf') as pdf:
            read = PdfReader(pdf)
            page = read.pages[0]
            text = page.extract_text()
            assert 'Пример pdf' in text


def test_csv_read():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('username.csv') as csv_mar:
            content = csv_mar.read().decode(
                'utf-8-sig')  # читаем содержимое файла и декодируем его если в файле есть символы не из английского алфавита
            csvreader = list(csv.reader(content.splitlines()))  # читаем содержимое файла и преобразуем его в список
            second_row = csvreader[1]  # получаем вторую строку
            assert ['booker12;9012;Rachel;Booker'] == second_row


def test_xlsx_read():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('file_xlsx.xlsx') as xlsx:
            wb = load_workbook(filename=xlsx)
            sheet = wb.active
            value = sheet.cell(row=2,column=3).value
            assert 'Коммерческий департамент' in value

